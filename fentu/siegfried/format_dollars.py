"""Restyle dollar columns of a siegfried workbook with thousands separators.

The ``roic_faustmann`` writer stores every value as raw text (e.g.
``1.330500e+11``). This module rewrites the dollar columns — EBIT, invested
capital, market cap, cash, debt, preferred equity, and net worth — as
whole dollars with a comma every three digits (e.g. ``133,050,000,000``),
updating the workbook in place. Ratios (roic, Faustmann ratio), the ticker,
and the rank/pick columns are left untouched. ``-`` placeholders and
unparseable text stay as-is.

Works on both .ods and .xlsx, from any folder. ``read_table`` strips the
commas back off, so re-running the picker on a formatted workbook still
parses cleanly.

Usage:
    uv run python -m fentu.siegfried.format_dollars data/ndx100_ticker_siegfried.ods
    uv run python -m fentu.siegfried.format_dollars data/ndx100_ticker_siegfried.ods --columns ebit net_worth
"""
import argparse
import math
import sys
from typing import List, Optional

import pandas as pd

DOLLAR_COLUMNS = (
    "ebit",
    "invested_capital",
    "market_cap",
    "cash",
    "debt",
    "preferred_equity",
    "net_worth",
)


def _cell_text(cell) -> str:
    from odf.text import P

    texts = [str(p) for p in cell.getElementsByType(P)]
    return texts[0] if texts else ""


def _set_text(cell, text: str) -> None:
    from odf.text import P

    for paragraph in cell.getElementsByType(P):
        cell.removeChild(paragraph)
    cell.addElement(P(text=text))


def _dollar_text(text: str) -> Optional[str]:
    """Comma-grouped whole dollars from a raw cell text, or None when not a finite dollar amount."""
    try:
        value = float(text.strip().replace(",", ""))
    except ValueError:
        return None
    if not math.isfinite(value):
        return None
    return f"{value:,.0f}"


def format_ods(path: str, columns: List[str]) -> int:
    """Rewrite the dollar columns' cell text in place; returns cells formatted."""
    from odf.opendocument import load
    from odf.table import Table, TableCell, TableRow

    wanted = [c.strip().lower() for c in columns]
    doc = load(path)
    table = doc.spreadsheet.getElementsByType(Table)[0]
    rows = table.getElementsByType(TableRow)
    header = [_cell_text(c) for c in rows[0].getElementsByType(TableCell)]
    indices = [i for i, name in enumerate(header) if name.strip().lower() in wanted]

    formatted = 0
    for row in rows[1:]:
        cells = row.getElementsByType(TableCell)
        for index in indices:
            if index >= len(cells):
                continue
            text = _dollar_text(_cell_text(cells[index]))
            if text is not None:
                _set_text(cells[index], text)
                formatted += 1
    doc.save(path)
    return formatted


def format_xlsx(path: str, columns: List[str]) -> int:
    """Give the dollar columns a #,##0 number format in place; returns cells styled."""
    import openpyxl

    wanted = [c.strip().lower() for c in columns]
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    header = {str(ws.cell(row=1, column=c).value).strip().lower(): c for c in range(1, ws.max_column + 1)}
    col_indices = [header[name] for name in wanted if name in header]

    formatted = 0
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in col_indices:
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, (int, float)) and math.isfinite(cell.value):
                cell.number_format = "#,##0"
                formatted += 1
    wb.save(path)
    return formatted


def format_workbook(path: str, columns: Optional[List[str]] = None) -> int:
    """Restyle the dollar columns of a siegfried workbook in place, by extension."""
    wanted = list(columns) if columns else list(DOLLAR_COLUMNS)
    if path.lower().endswith(".ods"):
        return format_ods(path, wanted)
    return format_xlsx(path, wanted)


def main(argv: Optional[list] = None) -> int:
    parser = argparse.ArgumentParser(description="Comma-group the dollar columns of a siegfried workbook")
    parser.add_argument("workbook", help="siegfried workbook (.ods or .xlsx)")
    parser.add_argument(
        "--columns", nargs="+", default=list(DOLLAR_COLUMNS), help="dollar columns to restyle (default: all)"
    )
    args = parser.parse_args(argv)

    formatted = format_workbook(args.workbook, args.columns)
    print(f"formatted {formatted} dollar cells in {args.workbook} (comma every three digits)")
    return 0


if __name__ == "__main__":
    sys.exit(main())