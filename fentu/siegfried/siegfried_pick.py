"""Pick the top slice of an index with Spitznagel's Chapter Ten screen.

From "Austrian Investing II: Siegfried" (The Dao of Capital, ch. 10), the
two-pronged screen — "neither of these screens in isolation is enough":

    1. High ROIC — Siegfrieds are operationally defined as firms realizing
       75 percent or higher ROIC (his own monthly toy screen even used ROIC
       above 100 percent). High ROIC flags roundabout, productive firms that
       keep reinvesting earnings.
    2. Low Faustmann ratio — among those survivors, the market-cap-to-net-
       worth laggards are the ones the market underappreciates ("the lowest
       Faustmann ratio firms among those with high ROIC").

The whole workbook is ranked in descending order of these criteria — ROIC
first (higher is better), then Faustmann ratio (lower is better) — with
non-positive net worth (negative Faustmann ratio) pushed below every
genuine candidate: a low ratio must mean underpricing, not a gutted balance
sheet. The top ``--top-n`` rows (default 10) are the picks.

Updates the workbook **in place** — no new file:

  * a "Siegfried rank" column (1..N over the whole universe),
  * a "Siegfried pick" column (YES on the top ``--top-n``),
  * a "Reason" column briefly stating why each row ranks where it does,
  * gold highlighting on the picked rows,
  * rows re-sorted picks-first (rank ascending),
  * the derived Faustmann ratio (column J) as a live formula for
    LibreOffice, ``=E/I``, so it recalculates when you edit the raw market
    cap / net-worth inputs. (.ods only — .xlsx keeps literal values so
    ``read_table`` still parses numerics.) The ROIC column stays a literal
    value: it is the rolling 10-year median from ``roic_faustmann``, not a
    single-year ``=B/C`` quotient.

Running it twice is safe: existing rank/pick/reason columns are reused, not
duplicated. Spitznagel ignores financials/banks in his own screen — do that
upstream when building the universe, as this table has no sector column.

Usage:
    uv run python -m fentu.siegfried.siegfried_pick data/ndx100_ticker_siegfried.ods
    uv run python -m fentu.siegfried.siegfried_pick data/ndx100_ticker_siegfried.ods --top-n 10 --min-roic 0.75
"""
import argparse
import math
import sys
from typing import List, Optional

import pandas as pd

from fentu.siegfried.roic_faustmann import read_table

DEFAULT_TOP_N = 10
DEFAULT_MIN_ROIC = 0.75
RANK_COLUMN = "Siegfried rank"
PICK_COLUMN = "Siegfried pick"
REASON_COLUMN = "Reason"
GOLD_STYLE = "siegfried_pick"
GOLD_BACKGROUND = "#FFD966"
GOLD_COLOR = "#7F6000"

FAUSTMANN_COLUMN = 9  # J: derived Faustmann ratio


def rank_siegfrieds(table: pd.DataFrame, min_roic: float = DEFAULT_MIN_ROIC) -> pd.DataFrame:
    """The whole universe ranked by Spitznagel's criteria, best first.

    Positive net worth first, then ROIC descending (higher is better), then
    Faustmann ratio ascending (lower is better); rank 1 = best candidate.
    ``min_roic`` only flavors the Reason text — the ordering never excludes
    a row, so the whole file gets ranked.
    """
    frame = table.copy()
    frame["_positive_net_worth"] = frame["faustmann_ratio"] > 0
    frame["_faustmann"] = frame["faustmann_ratio"].where(frame["_positive_net_worth"], float("inf"))
    ranked = frame.sort_values(
        ["_positive_net_worth", "roic", "_faustmann"],
        ascending=[False, False, True],
        na_position="last",
    )
    ranked = ranked.drop(columns=["_positive_net_worth", "_faustmann"]).reset_index(drop=True)
    ranked["rank"] = range(1, len(ranked) + 1)
    return ranked


def pick_siegfrieds(table: pd.DataFrame, top_n: int = DEFAULT_TOP_N) -> pd.DataFrame:
    """Top ``top_n`` of the ranked universe — the highlighted, invested-in slice."""
    return rank_siegfrieds(table).head(max(0, top_n))


def reason_for(roic: Optional[float], faustmann: Optional[float], min_roic: float, picked: bool) -> str:
    """One-line reason for a row's standing under the two-pronged screen."""
    bar = f"{min_roic:.0%}"
    if roic is None or pd.isna(roic):
        return "insufficient ROIC history (<10y)"
    if faustmann is None or pd.isna(faustmann):
        return "missing Faustmann data"
    if faustmann <= 0:
        return "non-positive net worth — not a low-Faustmann pick"
    above = roic >= min_roic
    status = "above" if above else "below"
    if picked:
        return f"ROIC {roic:.0%} {status} {bar} bar, low Faustmann {faustmann:.1f}"
    return f"ROIC {roic:.0%} {status} {bar} bar"


def annotate(table: pd.DataFrame, top_n: int, min_roic: float) -> List[dict]:
    """Per-row annotations (rank, picked, reason) in final row order."""
    ranked = rank_siegfrieds(table, min_roic=min_roic)
    out = []
    for _, row in ranked.iterrows():
        rank = int(row["rank"])
        out.append(
            {
                "ticker": row["ticker"],
                "rank": rank,
                "picked": rank <= top_n,
                "reason": reason_for(row["roic"], row["faustmann_ratio"], min_roic, rank <= top_n),
            }
        )
    return out


def _cell(text: str, style_name: Optional[str] = None):
    from odf.table import TableCell
    from odf.text import P

    cell = TableCell(stylename=style_name) if style_name else TableCell()
    cell.addElement(P(text=text))
    return cell


def _row_texts(row) -> list:
    from odf.table import TableCell
    from odf.text import P

    return [str(p) for c in row.getElementsByType(TableCell) for p in c.getElementsByType(P)]


def _cell_text(cell) -> str:
    from odf.text import P

    texts = [str(p) for p in cell.getElementsByType(P)]
    return texts[0] if texts else ""


def _set_text(cell, text: str) -> None:
    from odf.text import P

    for paragraph in cell.getElementsByType(P):
        cell.removeChild(paragraph)
    cell.addElement(P(text=text))


def _set_formula(cell, ref_a: str, ref_b: str, divisor_text: str) -> None:
    """Turn a cell into ``of:=ref_a/ref_b`` when the divisor is a finite number.

    The cached text stays, so ``read_table`` still parses the number;
    LibreOffice recalcs from the formula on open.
    """
    try:
        divisor = float(divisor_text.strip().replace(",", ""))
    except ValueError:
        return
    if not math.isfinite(divisor):
        return
    cell.setAttribute("formula", f"of:={ref_a}/{ref_b}")
    cell.setAttribute("valuetype", "float")


def update_ods(path: str, annotations: List[dict], top_n: int) -> None:
    """In-place .ods update: rank/pick/reason columns, gold picks, picks-first order, live formulas."""
    from odf.opendocument import load
    from odf.style import Style, TableCellProperties, TextProperties
    from odf.table import Table, TableCell, TableRow

    doc = load(path)
    table = doc.spreadsheet.getElementsByType(Table)[0]
    rows = table.getElementsByType(TableRow)

    existing = doc.automaticstyles.getElementsByType(Style)
    if not any(s.getAttribute("name") == GOLD_STYLE for s in existing):
        gold = Style(name=GOLD_STYLE, family="table-cell")
        gold.addElement(TableCellProperties(backgroundcolor=GOLD_BACKGROUND))
        gold.addElement(TextProperties(color=GOLD_COLOR, fontweight="bold"))
        doc.automaticstyles.addElement(gold)

    header = _row_texts(rows[0])
    rank_idx = header.index(RANK_COLUMN) if RANK_COLUMN in header else len(header)
    if RANK_COLUMN not in header:
        rows[0].addElement(_cell(RANK_COLUMN, "header"))
    if PICK_COLUMN not in header:
        rows[0].addElement(_cell(PICK_COLUMN, "header"))
    if REASON_COLUMN not in header:
        rows[0].addElement(_cell(REASON_COLUMN, "header"))

    by_ticker = {a["ticker"]: a for a in annotations}
    annotated = []
    for position, row in enumerate(rows[1:]):
        texts = _row_texts(row)
        ticker = texts[0] if texts else ""
        info = by_ticker.get(ticker)
        rank = info["rank"] if info else None
        picked = bool(info["picked"]) if info else False
        reason = info["reason"] if info else ""
        cells = row.getElementsByType(TableCell)
        while len(cells) <= rank_idx + 2:
            row.addElement(_cell("-"))
            cells = row.getElementsByType(TableCell)
        _set_text(cells[rank_idx], str(rank) if rank else "-")
        _set_text(cells[rank_idx + 1], "YES" if picked else "-")
        _set_text(cells[rank_idx + 2], reason or "-")
        if picked:
            for cell in row.getElementsByType(TableCell):
                cell.setAttribute("stylename", GOLD_STYLE)
        annotated.append((rank or 10**9, position, row))

    ordered = sorted(annotated, key=lambda item: (item[0], item[1]))
    for row in rows[1:]:
        table.removeChild(row)
    for _, _, row in ordered:
        table.addElement(row)

    for sheet_row, (_, _, row) in enumerate(ordered, start=2):
        cells = row.getElementsByType(TableCell)
        if len(cells) > FAUSTMANN_COLUMN:
            _set_formula(cells[FAUSTMANN_COLUMN], f"E{sheet_row}", f"I{sheet_row}", _cell_text(cells[8]))
    doc.save(path)


def update_xlsx(path: str, annotations: List[dict], top_n: int) -> None:
    """In-place .xlsx update: rank/pick/reason columns, gold picks, picks-first order (literal values)."""
    from copy import copy

    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    existing = [headers.get(name) for name in (RANK_COLUMN, PICK_COLUMN, REASON_COLUMN)]
    next_free = max(c for c in existing if c is not None) + 1 if any(existing) else ws.max_column + 1
    rank_col, pick_col, reason_col = existing
    if rank_col is None:
        rank_col = next_free
        next_free += 1
        ws.cell(row=1, column=rank_col, value=RANK_COLUMN)
    if pick_col is None:
        pick_col = next_free
        next_free += 1
        ws.cell(row=1, column=pick_col, value=PICK_COLUMN)
    if reason_col is None:
        reason_col = next_free
        ws.cell(row=1, column=reason_col, value=REASON_COLUMN)

    gold = PatternFill("solid", fgColor="FFD966")
    by_ticker = {a["ticker"]: a for a in annotations}
    snapshots = []
    for row_idx in range(2, ws.max_row + 1):
        ticker = str(ws.cell(row=row_idx, column=1).value or "").strip()
        info = by_ticker.get(ticker)
        rank = info["rank"] if info else None
        picked = bool(info["picked"]) if info else False
        ws.cell(row=row_idx, column=rank_col, value=rank if rank else "-")
        ws.cell(row=row_idx, column=pick_col, value="YES" if picked else "-")
        ws.cell(row=row_idx, column=reason_col, value=(info["reason"] if info else "") or "-")
        if picked:
            for col_idx in range(1, reason_col + 1):
                ws.cell(row=row_idx, column=col_idx).fill = gold
            ws.cell(row=row_idx, column=1).font = Font(bold=True)
        cells = [
            (ws.cell(row=row_idx, column=c).value, copy(ws.cell(row=row_idx, column=c)._style))
            for c in range(1, reason_col + 1)
        ]
        snapshots.append((rank or 10**9, row_idx, cells))

    ordered = sorted(snapshots, key=lambda snap: (snap[0], snap[1]))
    for new_row, (_, _, cells) in enumerate(ordered, start=2):
        for col_idx, (value, style) in enumerate(cells, start=1):
            cell = ws.cell(row=new_row, column=col_idx, value=value)
            cell._style = style
    wb.save(path)


def update_workbook(path: str, annotations: List[dict], top_n: int) -> None:
    """In-place update of a siegfried workbook (.ods or .xlsx), by extension."""
    if path.lower().endswith(".ods"):
        update_ods(path, annotations, top_n)
    else:
        update_xlsx(path, annotations, top_n)


def _fmt_pct(value: Optional[float]) -> str:
    return "-" if value is None or pd.isna(value) else f"{value * 100:.1f}%"


def _fmt_ratio(value: Optional[float]) -> str:
    return "-" if value is None or pd.isna(value) else f"{value:.2f}"


def main(argv: Optional[list] = None) -> int:
    parser = argparse.ArgumentParser(
        description="Siegfried picker: rank the whole workbook by Spitznagel's ch. 10 criteria, pick the top slice"
    )
    parser.add_argument("workbook", help="siegfried workbook (.ods or .xlsx) from roic_faustmann")
    parser.add_argument("--top-n", type=int, default=DEFAULT_TOP_N, help="number of picks to highlight (default 10)")
    parser.add_argument("--min-roic", type=float, default=DEFAULT_MIN_ROIC, help="ROIC bar for the reason text (default 0.75)")
    args = parser.parse_args(argv)

    table = read_table(args.workbook)
    annotations = annotate(table, top_n=args.top_n, min_roic=args.min_roic)
    picks = [a for a in annotations if a["picked"]]

    print(f"universe : {len(table)} tickers, ranked by Spitznagel's criteria (ROIC desc, Faustmann asc)")
    print(f"picks    : top {args.top_n} highlighted")
    for a in picks:
        row = table[table["ticker"] == a["ticker"]].iloc[0]
        print(
            f"  {a['rank']:>3}. {a['ticker']:<6} ROIC {_fmt_pct(row['roic']):>7}"
            f"  Faustmann {_fmt_ratio(row['faustmann_ratio']):>8}  {a['reason']}"
        )

    update_workbook(args.workbook, annotations, args.top_n)
    print(
        f"updated  : {args.workbook} (rank/pick/reason columns, gold top {args.top_n},"
        " picks first, J = E/I live formula)"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())