"""All-seven-traits CANSLIM report over the pharma/bio tables (O'Neil).

Schwager, Market Wizards (2006), O'Neil interview: "At any given time, less
than 2 percent of the stocks in the entire market will fit the CANSLIM
formula." This module runs criteria C, A, N, S, L, I per ticker (M once — it
gates the whole market) over the union of tickers in
``data/pharma_bio_canslim_c_screen.ods`` and
``data/biotech_biopharm_investment_checklist.xlsx``:

    1. appends per-criterion columns to both tables, and
    2. highlights (gold) the top 2 names by O'Neil's own ranking — letters
       passed, tie-broken by current-quarter EPS growth (C, "our first
       basic rule") then 12-month relative strength (L, "pick the leading
       stocks") — and prints the why for each pick.

Per-ticker results are cached (JSON) so an interrupted run resumes.

Usage:
    uv run python -m fentu.canslim.seven_traits_report
    uv run python -m fentu.canslim.seven_traits_report --workers 6
"""
import argparse
import json
import re
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from odf.opendocument import load
from odf.style import Style, TableCellProperties, TableColumnProperties, TextProperties
from odf.table import Table, TableCell, TableColumn, TableRow
from odf.text import P

from fentu.canslim.annual_eps import screen_annual_eps
from fentu.canslim.current_eps import score_current_eps
from fentu.canslim.institutional_sponsorship import screen_institutional_sponsorship
from fentu.canslim.market_direction import dual_market_direction
from fentu.canslim.new_highs import fetch_price_history, score_new_high
from fentu.canslim.relative_strength import DEFAULT_MIN_RANK, rs_rank, twelve_month_return
from fentu.canslim.shares_outstanding import screen_shares_outstanding

ODS_PATH = "data/pharma_bio_canslim_c_screen.ods"
XLSX_PATH = "data/biotech_biopharm_investment_checklist.xlsx"
CACHE_PATH = Path("/tmp/opencode/canslim7_cache.json")
WORKERS = 8
L_CHUNK = 50
FETCH_ATTEMPTS = 3
XLSX_FIRST_DATA_ROW = 3
XLSX_LAST_DATA_ROW = 26
XLSX_LAST_COLUMN = 47
CRITERION_COLUMNS = (
    "A: 5y EPS CAGR", "A",
    "N: dist from 52w high", "N",
    "S: shares out", "S",
    "L: RS rank", "L",
    "I: inst held", "I",
    "M: dist days GSPC/XBI", "M",
    "O'Neil letters",
)


def _fault_isolated(ticker: str, leg: str, fn) -> dict:
    try:
        return fn(ticker)
    except Exception as exc:  # yfinance raises a zoo; a leg failure must not kill the run
        return {"passed": False, "reason": f"fetch_error: {type(exc).__name__}"}


def _score_annual(ticker: str) -> dict:
    r = screen_annual_eps(ticker)
    return {"passed": r.passed, "cagr": r.cagr, "reason": r.reason or ""}


def _score_new_high(ticker: str) -> dict:
    s = score_new_high(fetch_price_history(ticker))
    return {
        "passed": s.reason is None,
        "distance": s.distance_from_high,
        "days": s.days_since_high,
        "volume_confirmed": s.volume_confirmed,
        "reason": s.reason or "",
    }


def _score_shares(ticker: str) -> dict:
    r = screen_shares_outstanding(ticker)
    return {"passed": r.passed, "shares": r.shares, "reason": r.reason or ""}


def _score_sponsorship(ticker: str) -> dict:
    r = screen_institutional_sponsorship(ticker)
    return {"passed": r.passed, "held_pct": r.held_pct, "holders": r.holder_count, "reason": r.reason or ""}


def _score_one(ticker: str) -> dict:
    legs = {"A": _score_annual, "N": _score_new_high, "S": _score_shares, "I": _score_sponsorship}
    return {name: _fault_isolated(ticker, name, fn) for name, fn in legs.items()}


def _load_cache() -> dict:
    if CACHE_PATH.exists():
        return json.loads(CACHE_PATH.read_text())
    return {}


def _save_cache(cache: dict) -> None:
    CACHE_PATH.write_text(json.dumps(cache, indent=1))


def _needs_fetch(entry: Optional[dict]) -> bool:
    if entry is None:
        return True
    return any(str(leg.get("reason", "")).startswith("fetch_error") for leg in entry.values() if isinstance(leg, dict))


def collect_scores(tickers: List[str], cache: dict, workers: int = WORKERS) -> dict:
    """Per-ticker A/N/S/I verdicts, threaded and cached; fetch errors retried."""
    pending = [t for t in tickers if _needs_fetch(cache.get(t))]
    for attempt in range(FETCH_ATTEMPTS):
        if not pending:
            break
        print(f"fetch round {attempt + 1}: {len(pending)} tickers x A/N/S/I ...")
        with ThreadPoolExecutor(max_workers=workers) as pool:
            futures = {pool.submit(_score_one, t): t for t in pending}
            for future in as_completed(futures):
                cache[futures[future]] = future.result()
        _save_cache(cache)
        pending = [t for t in tickers if _needs_fetch(cache.get(t))]
    return cache


def fetch_universe_closes(tickers: List[str]):
    """Trailing-1y daily closes for the whole universe, chunked batch download."""
    import pandas as pd
    import yfinance as yf

    frames = []
    for start in range(0, len(tickers), L_CHUNK):
        chunk = tickers[start:start + L_CHUNK]
        try:
            closes = yf.download(chunk, period="1y", auto_adjust=False, progress=False)["Close"]
        except Exception:
            continue
        if isinstance(closes, pd.Series):
            closes = closes.to_frame()
        frames.append(closes)
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, axis=1)


def compute_l_ranks(closes) -> Dict[str, dict]:
    """O'Neil's L: percentile rank of the 12-month return within the universe."""
    returns = {
        ticker: twelve_month_return([float(v) for v in closes[ticker].dropna()])
        for ticker in closes.columns
    }
    valid = {t: r for t, r in returns.items() if r is not None}
    return {
        ticker: {"ret": ret, "rank": rs_rank(list(valid.values()), ret), "passed": False}
        for ticker, ret in valid.items()
    } | {ticker: {"ret": None, "rank": None, "passed": False} for ticker, ret in returns.items() if ret is None}


def compute_m() -> dict:
    """O'Neil's M, once: the dual market+sector gate applies to every name."""
    dual = dual_market_direction()
    return {
        "passed": dual.passed,
        "market_dist_days": dual.market.distribution_days,
        "sector_dist_days": dual.sector.distribution_days,
        "market_reason": dual.market.reason,
        "sector_reason": dual.sector.reason,
    }


def ods_tickers(path: str) -> List[Tuple[str, str, str]]:
    """(ticker, criterion-C verdict, growth text) per data row, in table order."""
    doc = load(path)
    rows = doc.spreadsheet.getElementsByType(TableRow)[1:]
    out = []
    for row in rows:
        texts = [str(p) for c in row.getElementsByType(TableCell) for p in c.getElementsByType(P)]
        if texts and texts[0]:
            out.append((texts[0], texts[7] if len(texts) > 7 else "", texts[6] if len(texts) > 6 else ""))
    return out


def xlsx_tickers(path: str) -> List[Tuple[str, int, Optional[float], Optional[float]]]:
    """(ticker, row, current EPS, prior-year EPS) from the checklist."""
    import openpyxl

    ws = openpyxl.load_workbook(path)["Checklist"]
    out = []
    for row in range(XLSX_FIRST_DATA_ROW, XLSX_LAST_DATA_ROW + 1):
        ticker = ws.cell(row=row, column=1).value
        if ticker:
            out.append((str(ticker), row, ws.cell(row=row, column=25).value, ws.cell(row=row, column=26).value))
    return out


def parse_growth(text: Optional[str]) -> Optional[float]:
    if not text:
        return None
    match = re.search(r"([+-]?\d+(?:\.\d+)?)", str(text).replace("nan", ""))
    return float(match.group(1)) / 100.0 if match else None


def _xlsx_growth(value) -> Optional[float]:
    """Growth from the checklist's mixed column: numeric decimals or '+38.8%' / 'n/a ...' text."""
    if isinstance(value, (int, float)):
        return float(value)
    return parse_growth(value)


def _parse_rank(text: Optional[str]) -> Optional[float]:
    if text is None:
        return None
    match = re.search(r"(\d+(?:\.\d+)?)", str(text))
    return float(match.group(1)) if match else None


def _parse_letters(text: Optional[str]) -> int:
    if text is None:
        return 0
    match = re.match(r"(\d+)\s*/", str(text))
    return int(match.group(1)) if match else 0


def letter_count(c_passed: bool, scores: dict, l: dict, m: dict) -> int:
    total = int(c_passed)
    total += sum(int(scores[leg]["passed"]) for leg in ("A", "N", "S", "I"))
    total += int(l["passed"]) + int(m["passed"])
    return total


def oneil_sort_key(entry: dict) -> tuple:
    """O'Neil ranking: letters first, then current earnings (C), then leadership (L)."""
    rank = entry["l"]["rank"]
    return (entry["letters"], entry["c_growth"] or -1.0, rank if rank is not None else -1.0)


def build_entries(payloads: List[dict], l_ranks: dict, m: dict) -> List[dict]:
    """One scoring record per payload with everything the tables need, O'Neil-ranked."""
    entries = []
    for payload in payloads:
        ticker, scores = payload["ticker"], payload["scores"]
        l = l_ranks.get(ticker, {"ret": None, "rank": None, "passed": False})
        entries.append({
            "ticker": ticker,
            "row": payload["row"],
            "scores": scores,
            "l": l,
            "m": m,
            "c_passed": payload["c_passed"],
            "c_growth": payload["c_growth"],
            "letters": letter_count(payload["c_passed"], scores, l, m),
        })
    return sorted(entries, key=oneil_sort_key, reverse=True)


def _ods_row_payload(ticker: str, row: int, verdict: str, growth_text: str, cache: dict) -> dict:
    return {
        "ticker": ticker,
        "row": row,
        "scores": cache.get(ticker, {}),
        "c_passed": verdict == "PASS",
        "c_growth": parse_growth(growth_text),
    }


def _fmt_pct(value: Optional[float], digits: int = 1) -> str:
    return "n/a" if value is None else f"{value * 100:.{digits}f}%"


def _fmt_shares(value: Optional[float]) -> str:
    return "n/a" if value is None else f"{value / 1e6:.1f}M"


def _verdict(passed: bool) -> str:
    return "Yes" if passed else "No"


def criterion_cells(entry: dict) -> List[str]:
    """Values for the per-criterion columns, in CRITERION_COLUMNS order."""
    s, l, m = entry["scores"], entry["l"], entry["m"]
    return [
        _fmt_pct(s["A"]["cagr"]),
        _verdict(s["A"]["passed"]),
        _fmt_pct(s["N"]["distance"]),
        _verdict(s["N"]["passed"]),
        _fmt_shares(s["S"]["shares"]),
        _verdict(s["S"]["passed"]),
        "n/a" if l["rank"] is None else f"{l['rank']:.1f}",
        _verdict(l["passed"]),
        _fmt_pct(s["I"]["held_pct"]),
        _verdict(s["I"]["passed"]),
        f"{m['market_dist_days']} / {m['sector_dist_days']}",
        _verdict(m["passed"]),
        f"{entry['letters']}/7",
    ]


def _cell(text: str, style_name: Optional[str] = None) -> TableCell:
    cell = TableCell(stylename=style_name) if style_name else TableCell()
    cell.addElement(P(text=text))
    return cell


def update_ods(path: str, entries: List[dict], top2: List[str]) -> None:
    """Append the seven-traits columns to the criterion-C .ods; gold-highlight top 2."""
    doc = load(path)
    table = doc.spreadsheet.getElementsByType(Table)[0]
    rows = table.getElementsByType(TableRow)
    by_ticker = {e["ticker"]: e for e in entries}

    widths = ("3.0cm", "1.6cm") * 6 + ("3.0cm", "2.2cm")
    for column, width in enumerate(widths):
        col_style = Style(name=f"c7col{column}", family="table-column")
        col_style.addElement(TableColumnProperties(columnwidth=width))
        doc.automaticstyles.addElement(col_style)
        table.addElement(TableColumn(stylename=col_style))

    top_style = Style(name="oneil_top", family="table-cell")
    top_style.addElement(TableCellProperties(backgroundcolor="#FFD966"))
    top_style.addElement(TextProperties(color="#7F6000", fontweight="bold"))
    doc.automaticstyles.addElement(top_style)

    header = rows[0]
    for title in CRITERION_COLUMNS:
        header.addElement(_cell(title, "header"))

    for row in rows[1:]:
        cells = row.getElementsByType(TableCell)
        texts = [str(p) for c in cells for p in c.getElementsByType(P)]
        entry = by_ticker.get(texts[0]) if texts else None
        if entry is None:
            continue
        row_style = cells[0].getAttribute("stylename")
        for value in criterion_cells(entry):
            row.addElement(_cell(value, row_style))
        if entry["ticker"] in top2:
            for cell in row.getElementsByType(TableCell):
                cell.setAttribute("stylename", "oneil_top")

    doc.save(path)


def update_xlsx(path: str, entries: List[dict], top2: List[str]) -> None:
    """Insert the S/L/I (+ letters) columns before the Trading section; gold top 2."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.load_workbook(path)
    ws = wb["Checklist"]
    ws.insert_cols(34, 7)
    ws.unmerge_cells("AB1:AG1")
    ws.merge_cells("AB1:AN1")

    headers = [
        "Criterion S — shares outstanding",
        "Criterion S — verdict (Yes/No)",
        "Criterion L — 12-mo RS rank (pharma/bio universe)",
        "Criterion L — verdict (Yes/No)",
        "Criterion I — institutional held % (holders)",
        "Criterion I — verdict (Yes/No)",
        "O'Neil letters passed (of 7)",
    ]
    for offset, title in enumerate(headers):
        ws.cell(row=2, column=34 + offset, value=title)

    by_ticker = {e["ticker"]: e for e in entries}
    for entry in entries:
        row = entry["row"]
        s, l = entry["scores"], entry["l"]
        held = "n/a" if s["I"]["held_pct"] is None else f"{s['I']['held_pct'] * 100:.1f}% ({s['I']['holders']})"
        values = [
            _fmt_shares(s["S"]["shares"]),
            _verdict(s["S"]["passed"]),
            "n/a" if l["rank"] is None else f"{l['rank']:.1f}",
            _verdict(l["passed"]),
            held,
            _verdict(s["I"]["passed"]),
            entry["letters"],
        ]
        for offset, value in enumerate(values):
            ws.cell(row=row, column=34 + offset, value=value)

    gold = PatternFill("solid", fgColor="FFD966")
    for ticker in top2:
        entry = by_ticker.get(ticker)
        if entry is None:
            continue
        for column in range(1, 48):
            ws.cell(row=entry["row"], column=column).fill = gold
        ws.cell(row=entry["row"], column=1).font = Font(bold=True)
    wb.save(path)


def _ods_row_texts(row) -> List[str]:
    return [str(p) for c in row.getElementsByType(TableCell) for p in c.getElementsByType(P)]


def reorder_ods(path: str) -> List[str]:
    """Re-sort .ods data rows by the O'Neil ranking already stored in the columns.

    Descending: letters passed, then current-quarter EPS growth (C), then
    12-month relative-strength rank (L). Styles (including the gold top-2)
    travel with their rows. Returns the new ticker order.
    """
    doc = load(path)
    table = doc.spreadsheet.getElementsByType(Table)[0]
    rows = table.getElementsByType(TableRow)
    data = rows[1:]

    def sort_key(row) -> tuple:
        texts = _ods_row_texts(row)
        growth = parse_growth(texts[6]) if len(texts) > 6 else None
        rank = _parse_rank(texts[15]) if len(texts) > 15 else None
        letters = _parse_letters(texts[21]) if len(texts) > 21 else 0
        return (letters, growth if growth is not None else -1.0, rank if rank is not None else -1.0)

    ordered = sorted(data, key=sort_key, reverse=True)
    for row in data:
        table.removeChild(row)
    for row in ordered:
        table.addElement(row)
    doc.save(path)
    return [_ods_row_texts(row)[0] for row in ordered]


def _xlsx_remap_formula(formula: str, old_row: int, new_row: int) -> str:
    """Point a row's own-referencing formula at its new row (range refs untouched)."""
    return re.sub(
        rf"(?<!\$)\b([A-Z]{{1,2}}){old_row}\b",
        rf"\g<1>{new_row}",
        formula,
    )


def reorder_xlsx(path: str) -> List[str]:
    """Re-sort the checklist's data rows by the stored O'Neil ranking (descending).

    Moves values, formulas (own-row references remapped), and cell styles;
    the gold top-2 highlight travels with its rows. Returns the new order.
    """
    from copy import copy

    import openpyxl

    wb = openpyxl.load_workbook(path)
    ws = wb["Checklist"]

    snapshots = []
    for old_row in range(XLSX_FIRST_DATA_ROW, XLSX_LAST_DATA_ROW + 1):
        if not ws.cell(row=old_row, column=1).value:
            continue
        cells = [ws.cell(row=old_row, column=col) for col in range(1, XLSX_LAST_COLUMN + 1)]
        growth = _xlsx_growth(ws.cell(row=old_row, column=27).value)
        rank = _parse_rank(ws.cell(row=old_row, column=36).value)
        letters = int(ws.cell(row=old_row, column=40).value or 0)
        snapshots.append({
            "old_row": old_row,
            "key": (letters, growth if growth is not None else -1.0, rank if rank is not None else -1.0),
            "cells": [(cell.value, copy(cell._style)) for cell in cells],
        })
    ordered = sorted(snapshots, key=lambda s: s["key"], reverse=True)

    for new_idx, snap in enumerate(ordered):
        new_row = XLSX_FIRST_DATA_ROW + new_idx
        for col_idx, (value, style) in enumerate(snap["cells"], start=1):
            cell = ws.cell(row=new_row, column=col_idx)
            if isinstance(value, str) and value.startswith("="):
                value = _xlsx_remap_formula(value, snap["old_row"], new_row)
            cell.value = value
            cell._style = style
    wb.save(path)
    return [str(snap["cells"][0][0]) for snap in ordered]


def why(entry: dict) -> str:
    """O'Neil's own voice, citing the numbers that drove the pick."""
    s, l = entry["scores"], entry["l"]
    bits = [
        f"{entry['letters']}/7 CANSLIM letters",
        f"C: quarterly EPS {_fmt_pct(entry['c_growth'])} YoY" if entry["c_growth"] is not None else "C: n/a",
        f"A: 5y CAGR {_fmt_pct(s['A']['cagr'])}",
        f"N: {_fmt_pct(s['N']['distance'])} off the 52-wk closing high",
        f"S: {_fmt_shares(s['S']['shares'])} shares out",
        f"L: RS rank {l['rank']:.1f}" if l["rank"] is not None else "L: n/a",
        f"I: {_fmt_pct(s['I']['held_pct'])} institutionally held",
    ]
    return f"{entry['ticker']}: " + "; ".join(bits)


def main(argv: Optional[list] = None) -> int:
    parser = argparse.ArgumentParser(description="Seven-traits CANSLIM report over the two pharma/bio tables")
    parser.add_argument("--workers", type=int, default=WORKERS, help="fetch threads (default 8)")
    parser.add_argument("--ods", default=ODS_PATH, help=".ods table to update")
    parser.add_argument("--xlsx", default=XLSX_PATH, help=".xlsx table to update")
    parser.add_argument("--dry-run", action="store_true", help="score and print only; do not write files")
    parser.add_argument("--reorder", action="store_true", help="only re-sort both tables by the stored O'Neil ranking (no network)")
    args = parser.parse_args(argv)

    if args.reorder:
        ods_order = reorder_ods(args.ods)
        xlsx_order = reorder_xlsx(args.xlsx)
        print(f"reordered (most -> least recommended):")
        print(f"  ods : {', '.join(ods_order[:10])} ... ({len(ods_order)} rows)")
        print(f"  xlsx: {', '.join(xlsx_order)}")
        return 0

    ods_rows = ods_tickers(args.ods)
    xlsx_rows = xlsx_tickers(args.xlsx)
    print(f"ods: {len(ods_rows)} tickers | xlsx: {len(xlsx_rows)} tickers")

    universe = list(dict.fromkeys([t for t, *_ in ods_rows] + [t for t, *_ in xlsx_rows]))
    cache = collect_scores(universe, _load_cache(), args.workers)
    _save_cache(cache)

    print("ranking universe on 12-month relative strength ...")
    closes = fetch_universe_closes(universe)
    l_ranks = compute_l_ranks(closes)
    for ticker, entry in l_ranks.items():
        entry["passed"] = entry["rank"] is not None and entry["rank"] >= DEFAULT_MIN_RANK - 1e-9

    print("scoring criterion M (market direction, dual gate) ...")
    m = compute_m()
    print(f"M: passed={m['passed']} dist days GSPC {m['market_dist_days']} / XBI {m['sector_dist_days']}")

    ods_entries = build_entries(
        [_ods_row_payload(t, i + 2, verdict, growth, cache)
         for i, (t, verdict, growth) in enumerate(ods_rows)],
        l_ranks, m,
    )

    xlsx_entries = build_entries(
        [_xlsx_row_payload(t, row, cur, prior, cache) for t, row, cur, prior in xlsx_rows],
        l_ranks, m,
    )

    for label, entries in (("ODS (criterion-C universe)", ods_entries), ("XLSX (checklist)", xlsx_entries)):
        top2 = [e["ticker"] for e in entries[:2]]
        print(f"\n== {label}: top 2 by O'Neil ==")
        for entry in entries[:2]:
            print(" " + why(entry))
        if not args.dry_run:
            if label.startswith("ODS"):
                update_ods(args.ods, entries, top2)
            else:
                update_xlsx(args.xlsx, entries, top2)
            print(f" updated: {args.ods if label.startswith('ODS') else args.xlsx}")
    return 0


def _xlsx_row_payload(ticker: str, row: int, current, prior, cache: dict) -> dict:
    try:
        result = score_current_eps(float(current), float(prior))
        growth = None if result.growth != result.growth else result.growth  # NaN guard (negative base)
        return {"ticker": ticker, "row": row, "scores": cache[ticker], "c_passed": result.passed, "c_growth": growth}
    except (TypeError, ValueError):  # blank or 'n/a ...' text cells
        return {"ticker": ticker, "row": row, "scores": cache[ticker], "c_passed": False, "c_growth": None}


if __name__ == "__main__":
    sys.exit(main())
