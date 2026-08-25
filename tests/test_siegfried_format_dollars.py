"""
Dollar restyling of siegfried workbooks — comma every three digits.

Covered behaviors:
    1. .ods: dollar cells are rewritten in place (e.g. 1.33e+11 ->
       133,050,000,000); ratios, ticker, and "-" placeholders are untouched.
    2. .xlsx: dollar cells get the #,##0 number format; text stays text.
    3. --columns limits the restyle to the named columns only.
    4. read_table strips the commas back off, so a formatted workbook still
       feeds the picker with clean numerics.

Network-free: tables are built inline and round-tripped through tmp files.
"""
import pandas as pd
import pytest

from fentu.siegfried.format_dollars import DOLLAR_COLUMNS, format_workbook
from fentu.siegfried.roic_faustmann import OUTPUT_COLUMNS, read_table, write_table
from fentu.siegfried.siegfried_pick import pick_siegfrieds


def _table():
    return pd.DataFrame(
        [
            {
                "ticker": "ABC",
                "ebit": 1_330_500_000_000.0,
                "invested_capital": 1_723_900_000_000.0,
                "roic": 0.771797,
                "market_cap": 4_529_158_000_000.0,
                "cash": 54_697_000_000.0,
                "debt": 98_657_000_000.0,
                "preferred_equity": None,
                "net_worth": 128_430_000_000.0,
                "faustmann_ratio": 35.265575,
            },
            {
                "ticker": "DEF",
                "ebit": None,
                "invested_capital": 5_000_000_000.0,
                "roic": 0.5,
                "market_cap": 10_000_000_000.0,
                "cash": 1_000_000.0,
                "debt": 0.0,
                "preferred_equity": 0.0,
                "net_worth": 5_001_000_000.0,
                "faustmann_ratio": 2.0,
            },
        ]
    ).reindex(columns=OUTPUT_COLUMNS)


def _cell_texts(rows, index):
    from odf.table import TableCell
    from odf.text import P

    return [
        [str(p) for c in row.getElementsByType(TableCell) for p in c.getElementsByType(P)][index]
        for row in rows[1:]
    ]


def _ods_rows(path):
    from odf.opendocument import load
    from odf.table import Table, TableRow

    doc = load(path)
    table = doc.spreadsheet.getElementsByType(Table)[0]
    return table.getElementsByType(TableRow)


def test_format_ods_commas_every_three_digits(tmp_path):
    path = str(tmp_path / "siegfried.ods")
    write_table(_table(), path)

    formatted = format_workbook(path)
    rows = _ods_rows(path)

    assert formatted == 12  # 7 dollar columns x 2 rows, minus the two NaN cells ("nan" text)
    assert _cell_texts(rows, 1)[0] == "1,330,500,000,000"
    assert _cell_texts(rows, 2)[0] == "1,723,900,000,000"
    assert _cell_texts(rows, 3)[0] == "0.771797"  # roic untouched
    assert _cell_texts(rows, 4)[0] == "4,529,158,000,000"
    assert _cell_texts(rows, 6)[0] == "98,657,000,000"
    assert _cell_texts(rows, 8)[0] == "128,430,000,000"
    assert _cell_texts(rows, 9)[0] == "35.265575"  # faustmann ratio untouched
    assert _cell_texts(rows, 1)[1] == "nan"  # missing EBIT placeholder untouched
    assert _cell_texts(rows, 5)[1] == "1,000,000"
    assert _cell_texts(rows, 7)[1] == "0"  # 0.0 preferred equity -> "0"


def test_format_ods_honors_columns_selection(tmp_path):
    path = str(tmp_path / "siegfried.ods")
    write_table(_table(), path)

    format_workbook(path, columns=["ebit"])

    rows = _ods_rows(path)
    assert _cell_texts(rows, 1)[0] == "1,330,500,000,000"
    assert _cell_texts(rows, 2)[0] == "1723900000000.0"  # invested_capital untouched


def test_format_xlsx_sets_number_format(tmp_path):
    import openpyxl

    path = str(tmp_path / "siegfried.xlsx")
    write_table(_table(), path)

    formatted = format_workbook(path)
    assert formatted == 12

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    assert ws.cell(row=2, column=2).number_format == "#,##0"  # ebit
    assert ws.cell(row=2, column=3).number_format == "#,##0"  # invested capital
    assert ws.cell(row=2, column=4).number_format != "#,##0"  # roic untouched
    assert ws.cell(row=2, column=4).value == pytest.approx(0.771797)
    assert ws.cell(row=2, column=9).number_format == "#,##0"  # net worth
    assert ws.cell(row=2, column=10).number_format != "#,##0"  # faustmann ratio untouched


@pytest.mark.parametrize("suffix", [".ods", ".xlsx"])
def test_read_table_strips_commas_after_formatting(tmp_path, suffix):
    table = _table()
    path = str(tmp_path / f"siegfried{suffix}")
    write_table(table, path)
    format_workbook(path)

    loaded = read_table(path)
    assert loaded["ebit"].tolist()[0] == pytest.approx(1_330_500_000_000.0)
    assert pd.isna(loaded["ebit"].tolist()[1])  # NaN placeholder survives as NaN
    assert loaded["net_worth"].tolist() == pytest.approx([128_430_000_000.0, 5_001_000_000.0])
    assert loaded["roic"].tolist() == pytest.approx([0.771797, 0.5])

    picks = pick_siegfrieds(loaded, top_n=2)
    assert list(picks["ticker"]) == ["ABC", "DEF"]  # picker still works on the restyled file